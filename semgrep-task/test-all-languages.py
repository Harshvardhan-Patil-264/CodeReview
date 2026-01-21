#!/usr/bin/env python3
"""
Multi-Language Test Script
Tests all vulnerable files to verify custom and community rules trigger correctly
"""

import subprocess
import json
import openpyxl
from pathlib import Path

def test_language(filename, language):
    """Test a single language file"""
    print(f"\n{'='*70}")
    print(f"Testing {language.upper()}: {filename}")
    print('='*70)
    
    # Run scanner
    print(f"Running scanner...")
    result = subprocess.run(
        ['python', 'auto-review.py', f'code/{filename}'],
        capture_output=True,
        text=True,
        encoding='utf-8'
    )
    
    if result.returncode != 0:
        print(f"❌ Scanner failed for {filename}")
        print(result.stdout)
        print(result.stderr)
        return None
    
    # Check Excel report
    report_name = f"reports/{Path(filename).stem}_Review.xlsx"
    
    if not Path(report_name).exists():
        print(f"❌ Report not found: {report_name}")
        return None
    
    # Analyze report
    wb = openpyxl.load_workbook(report_name)
    ws = wb.active
    
    total = ws.max_row - 1
    custom = 0
    community = 0
    rules = set()
    
    for row in range(2, ws.max_row + 1):
        rule_id = ws.cell(row, 4).value
        if rule_id:
            rules.add(rule_id)
            if 'community-rules' in rule_id:
                community += 1
            else:
                custom += 1
    
    print(f"\n✅ Results for {language}:")
    print(f"   Total Findings: {total}")
    print(f"   Custom Rules: {custom}")
    print(f"   Community Rules: {community}")
    print(f"   Unique Rules: {len(rules)}")
    
    return {
        'language': language,
        'total': total,
        'custom': custom,
        'community': community,
        'unique_rules': len(rules)
    }

def main():
    print("="*70)
    print("MULTI-LANGUAGE VULNERABILITY TESTING")
    print("="*70)
    print("Testing all vulnerable files to verify rule coverage...")
    
    tests = [
        ('test-vulnerable.go', 'Go'),
        ('test-vulnerable.js', 'JavaScript'),
        ('test-vulnerable.py', 'Python'),
        ('TestVulnerable.java', 'Java'),
        ('TestSpringBoot.java', 'Java Spring Boot')
    ]
    
    results = []
    for filename, language in tests:
        result = test_language(filename, language)
        if result:
            results.append(result)
    
    # Summary
    print(f"\n{'='*70}")
    print("SUMMARY - ALL LANGUAGES")
    print('='*70)
    print(f"{'Language':<15} {'Total':<10} {'Custom':<10} {'Community':<12} {'Rules':<10}")
    print('-'*70)
    
    for r in results:
        print(f"{r['language']:<15} {r['total']:<10} {r['custom']:<10} {r['community']:<12} {r['unique_rules']:<10}")
    
    total_findings = sum(r['total'] for r in results)
    total_custom = sum(r['custom'] for r in results)
    total_community = sum(r['community'] for r in results)
    
    print('-'*70)
    print(f"{'TOTAL':<15} {total_findings:<10} {total_custom:<10} {total_community:<12}")
    print('='*70)
    
    print(f"\n✅ Tested {len(results)} languages successfully!")
    print(f"✅ Total findings across all languages: {total_findings}")

if __name__ == "__main__":
    main()
